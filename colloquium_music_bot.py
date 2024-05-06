import telebot
from PIL import Image
from time import sleep
from random import randint
from difflib import SequenceMatcher
from math import ceil
import openpyxl

bot = telebot.TeleBot('TOKEN')
last = {}
list_mistakes = {}
answers = {}
hypothetical_mistakes = {}
persons = []
users = []
admin_ids = []
secret_word = 'SECRET'
admin_state = False


def checking_similarity(a, b):
    count_ns_a = count_ns_b = 0
    ns_a = ns_b = ''
    for i in a:
        if i in '0123456789' or i in ' ,;и':
            count_ns_a += 1
            if i not in ' ,;lи':
                ns_a += i
    for i in b:
        if i in '0123456789' or i in ' ,;и':
            count_ns_b += 1
            if i not in ' ,;lи':
                ns_b += i
    if count_ns_a == len(a) and count_ns_a == len(b):
        if ns_a == ns_b:
            return 100
        else:
            if a > b:
                return (1 - ceil(int(ns_b) / int(ns_a))) * 100
            else:
                return (1 - ceil(int(ns_a) / int(ns_b))) * 100
    else:
        return int(SequenceMatcher(None, a, b).ratio() * 100)


def read_questions():
    global persons
    path = 'questions.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    n_row = sheet_obj.max_row
    n_column = sheet_obj.max_column
    for column in range(2, n_column + 1):
        persons.append([sheet_obj.cell(row=1, column=column).value])
        for row in range(2, n_row, 2):
            question = sheet_obj.cell(row=row, column=column).value
            answer = str(sheet_obj.cell(row=row + 1, column=column).value)
            if question:
                persons[column - 2].append([question, answer])


read_questions()


@bot.message_handler(commands=['start'])
def start(message):
    user_id = message.chat.id
    if message.from_user.last_name:
        user = [user_id, message.from_user.username, message.from_user.first_name, message.from_user.last_name]
        if user not in users:
            users.append(user)
    else:
        user = [user_id, message.from_user.username, message.from_user.first_name]
        if user not in users:
            users.append(user)
    list_mistakes[user_id] = []
    answers[user_id] = 0
    hypothetical_mistakes[user_id] = []
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    composers = telebot.types.KeyboardButton('Композиторы')
    mistakes = telebot.types.KeyboardButton('Мои ошибки')
    reset = telebot.types.KeyboardButton('Сбросить мои ошибки')
    rate = telebot.types.KeyboardButton('Процент правильных ответов')
    list_commands = telebot.types.KeyboardButton('Помощь')
    # reboot = telebot.types.KeyboardButton('/start')  # отладочная кнопка. для работы добавить в строку ниже!
    markup.add(composers, mistakes, reset, rate, list_commands)
    bot.send_message(user_id,
                     'Добро пожаловать, {0.first_name}!\nЯ - <b>{1.first_name}</b>! Я помогу тебе подготовиться...'.format(
                         message.from_user, bot.get_me()), parse_mode='html', reply_markup=markup)
    bot.register_next_step_handler(message, main_menu)


@bot.message_handler(regexp='Пользователи')
def list_users(message):
    global users
    user_id = message.chat.id
    if user_id in admin_ids:
        if users:
            k = 1
            for user in users:
                user_full_name = ' '.join(user[2:])
                user_link = user[0]
                bot.send_message(
                    user_id, f'[{user_full_name}](tg://user?id={user_link})', parse_mode='MarkdownV2')
                k += 1
        else:
            bot.send_message(user_id, 'Пока нет ни одного пользователя...')
    else:
        bot.send_message(user_id, 'Прости, не могу тебе этого сказать, ты не мой создатель...')


@bot.message_handler(regexp='Рейтинг')
def get_rating(message):
    global users
    user_id = message.chat.id
    if user_id in admin_ids:
        if answers[user_id]:
            k = 1
            rating = sorted([(ceil(((answers[i] - len(list_mistakes[i])) / answers[i]) * 100)) for i in
                             list(list_mistakes.keys())], reverse=True)
            for user in users:
                user_full_name = ' '.join(user[2:])
                user_link = user[0]
                percent = ceil(((answers[user_link] - len(list_mistakes[user_link])) / answers[user_link]) * 100)
                place = rating.index(percent) + 1
                bot.send_message(user_id, f'{place} место:')
                bot.send_message(
                    user_id, f'[{user_full_name}](tg://user?id={user_link})', parse_mode='MarkdownV2')
                k += 1
        else:
            bot.send_message(user_id, 'Пока никто не отвечал на вопросы...')
    else:
        bot.send_message(user_id, 'Прости, не могу тебе этого сказать, ты не мой создатель...')


@bot.message_handler(regexp='Получить права администратора')
def append_admins(message):
    global admin_state
    user_id = message.chat.id
    if user_id in admin_ids:
        bot.send_message(user_id, 'Вы уже имеете права администратора')
    else:
        admin_state = True
        bot.send_message(user_id, 'Назовите кодовое слово, чтобы получить права администратора...')
        bot.register_next_step_handler(message, main_menu)


@bot.message_handler(content_types=['text'])
def main_menu(message):
    global admin_state
    main_inline = telebot.types.InlineKeyboardMarkup()
    user_id = message.chat.id
    if message.text == 'Композиторы':
        key_1 = telebot.types.InlineKeyboardButton(text='Ф. Куперен', callback_data=f'0;1;{user_id}')
        key_2 = telebot.types.InlineKeyboardButton(text='Ж. Ф. Рамо', callback_data=f'0;2;{user_id}')
        key_3 = telebot.types.InlineKeyboardButton(text='И. С. Бах', callback_data=f'0;3;{user_id}')
        key_4 = telebot.types.InlineKeyboardButton(text='Ф. Й. Гайдн', callback_data=f'0;4;{user_id}')
        key_5 = telebot.types.InlineKeyboardButton(text='В. А. Моцарт', callback_data=f'0;5;{user_id}')
        key_6 = telebot.types.InlineKeyboardButton(text='Л. ван Бетховен', callback_data=f'0;6;{user_id}')
        key_7 = telebot.types.InlineKeyboardButton(text='Р. Шуман', callback_data=f'0;7;{user_id}')
        key_8 = telebot.types.InlineKeyboardButton(text='Ф. Шопен', callback_data=f'0;8;{user_id}')
        key_9 = telebot.types.InlineKeyboardButton(text='Ф. Лист', callback_data=f'0;9;{user_id}')
        key_10 = telebot.types.InlineKeyboardButton(text='И. Брамс', callback_data=f'0;10;{user_id}')
        key_11 = telebot.types.InlineKeyboardButton(text='К. А. Дебюсси', callback_data=f'0;11;{user_id}')
        key_12 = telebot.types.InlineKeyboardButton(text='М. Равель', callback_data=f'0;12;{user_id}')
        key_13 = telebot.types.InlineKeyboardButton(text='Н. К. Метнер', callback_data=f'0;13;{user_id}')
        key_14 = telebot.types.InlineKeyboardButton(text='С. В. Рахманинов', callback_data=f'0;14;{user_id}')
        key_15 = telebot.types.InlineKeyboardButton(text='А. Н. Скрябин', callback_data=f'0;15;{user_id}')
        key_16 = telebot.types.InlineKeyboardButton(text='Д. Д. Шостакович', callback_data=f'0;16;{user_id}')
        key_17 = telebot.types.InlineKeyboardButton(text='С. С. Прокофьев', callback_data=f'0;17;{user_id}')
        key_18 = telebot.types.InlineKeyboardButton(text='Случайный',
                                                    callback_data=f'0;{randint(1, 17)};{user_id}')
        main_inline.add(key_1, key_2, key_3, key_4, key_5, key_6, key_7, key_8, key_9, key_10, key_11, key_12, key_13,
                        key_14, key_15, key_16, key_17, key_18)
        bot.send_message(user_id, text='Выберите композитора:', reply_markup=main_inline)
    elif message.text == 'Мои ошибки':
        if list_mistakes[user_id]:
            bot.send_message(user_id, text='Вот вопросы, в которых вы ошибались:', reply_markup=main_inline)
            for mistake in list_mistakes[user_id]:
                bot.send_message(user_id, text=f'Вопрос: {mistake[0]}\nВерный ответ: {mistake[1]}',
                                 reply_markup=main_inline)
        else:
            bot.send_message(user_id, text='Вы идеальны...', reply_markup=main_inline)
    elif message.text == 'Сбросить мои ошибки':
        bot.send_message(user_id, text='Слушаюсь и повинуюсь...', reply_markup=main_inline)
        list_mistakes[user_id] = []
        answers[user_id] = 0
    elif message.text == 'Процент правильных ответов':
        if answers[user_id]:
            percent = ceil(((answers[user_id] - len(list_mistakes[user_id])) / answers[user_id]) * 100)
            place = sorted([(ceil(((answers[i] - len(list_mistakes[i])) / answers[i]) * 100)) for i in
                            list(list_mistakes.keys())], reverse=True).index(percent) + 1
            bot.send_message(user_id, text=f'Вы ответили верно на {percent}% вопросов...\nВы {place} в рейтинге!')
        else:
            bot.send_message(user_id, text='Вы ещё не ответили ни на один вопрос...')
    elif message.text == 'Помощь':
        help_markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
        get_list_users = telebot.types.KeyboardButton('Пользователи')
        rating = telebot.types.KeyboardButton('Рейтинг')
        get_privileges = telebot.types.KeyboardButton('Получить права администратора')
        get_out = telebot.types.KeyboardButton('Выйти в главное меню')
        help_markup.add(get_list_users, rating, get_privileges, get_out)
        bot.send_message(user_id,
                         text='Теперь у тебя есть доступ к дополнительным командам!\nЕсли это не то, что ты хотел, мой создатель - @user_nuII!\nОбратись к нему...',
                         reply_markup=help_markup)
    elif message.text == secret_word:
        admin_state = False
        admin_ids.append(user_id)
        bot.send_message(user_id, 'Теперь ты администратор! У тебя есть доступ ко всем командам...')
    elif admin_state:
        admin_state = False
        bot.send_message(user_id, 'Увы, неверно...')
    elif message.text == 'Выйти в главное меню':
        markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
        composers = telebot.types.KeyboardButton('Композиторы')
        mistakes = telebot.types.KeyboardButton('Мои ошибки')
        reset = telebot.types.KeyboardButton('Сбросить мои ошибки')
        rate = telebot.types.KeyboardButton('Процент правильных ответов')
        list_commands = telebot.types.KeyboardButton('Помощь')
        markup.add(composers, mistakes, reset, rate, list_commands)
        bot.send_message(user_id, text='Слушаюсь и повинуюсь',
                         reply_markup=markup)
    else:
        bot.send_message(user_id, text='Я не настолько умён. Пожалуйста, задай другой вопрос...',
                         reply_markup=main_inline)


@bot.callback_query_handler(func=lambda call: True)
def callback_query_handler(call):
    global last
    flag, composer_id, user_id = map(int, call.data.split(';'))
    if flag == 2:
        del list_mistakes[user_id][-1]
        bot.send_message(user_id, text='Каюсь, ошибся. Твоя ошибка отменена...')
    elif flag == 1:
        person = persons[composer_id - 1]
        r = randint(1, len(person) - 1)
        theme = person[r]
        question = theme[0]
        right_answer = theme[1]
        composer_name = person[0]
        last[user_id] = [question, right_answer, composer_id, composer_name]
        bot.send_message(user_id, question)
        bot.register_next_step_handler_by_chat_id(user_id, quiz)
    elif flag == 3:
        list_mistakes[user_id].append(hypothetical_mistakes[user_id])
        bot.send_message(user_id, text='Каюсь, ошибся. Твоя ошибка сохранена...')
    else:
        im = Image.open(f'{composer_id}.jpg')
        person = persons[composer_id - 1]
        r = randint(1, len(person) - 1)
        theme = person[r]
        question = theme[0]
        right_answer = theme[1]
        composer_name = person[0]
        last[user_id] = [question, right_answer, composer_id, composer_name]
        bot.send_message(user_id, composer_name)
        bot.send_photo(user_id, im)
        sleep(3)
        bot.send_message(user_id, question)
        bot.register_next_step_handler_by_chat_id(user_id, quiz)


def quiz(message):
    user_id = message.chat.id
    right_answer = last[user_id][1]
    question = last[user_id][0]
    person_id = last[user_id][2]
    answers[user_id] += 1
    constant_inline = telebot.types.InlineKeyboardMarkup()
    key = telebot.types.InlineKeyboardButton(text=last[user_id][3],
                                             callback_data=f'1;{person_id};{user_id}')
    constant_inline.add(key)
    if message.text == right_answer:
        bot.send_message(user_id, 'Верно!')
    elif checking_similarity(message.text, right_answer) < 50:
        custom_inline = telebot.types.InlineKeyboardMarkup()
        key = telebot.types.InlineKeyboardButton(text='Мой ответ верный!',
                                                 callback_data=f'2;0;{user_id}')
        custom_inline.add(key)
        bot.send_message(user_id, f'Кажется, вы ответили неверно...\nВерный ответ: {right_answer}',
                         reply_markup=custom_inline)
        list_mistakes[user_id].append([question, right_answer])
        sleep(3)
    else:
        hypothetical_mistakes[user_id] = [question, right_answer]
        custom_inline = telebot.types.InlineKeyboardMarkup()
        key = telebot.types.InlineKeyboardButton(text='Мой ответ неверный...', callback_data=f'3;0;{user_id}')
        custom_inline.add(key)
        bot.send_message(user_id,
                         f'Сойдёт... Ваш ответ совпад с идеальным на {checking_similarity(message.text, right_answer)}%.\nИдеальный ответ: {right_answer}',
                         reply_markup=custom_inline)
        sleep(3)
    bot.send_message(user_id,
                     'Чтобы продолжить отвечать на вопросы про этого композитора, нажмите на кнопку ниже:',
                     reply_markup=constant_inline)


bot.infinity_polling()
